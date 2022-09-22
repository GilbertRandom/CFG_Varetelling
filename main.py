import openpyxl
import os
import re
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

for file in os.listdir("./"):
    if file.startswith("CFG Ansatte"):
        employee_sales = file
    elif file.startswith("CrossFit Grimstad"):
        customer_sales = file
    elif file.startswith("varetelling"):
        varetelling = file


goods = openpyxl.load_workbook("varer.xlsx")
sheet_goods = goods.active

sales_employee = openpyxl.load_workbook(employee_sales)
sheet_sales_employee = sales_employee.active

sales_customer = openpyxl.load_workbook(customer_sales)
sheet_sales_customer = sales_customer.active

varetelling = openpyxl.load_workbook(varetelling)
sheet_varetelling = varetelling.active

income_report = openpyxl.load_workbook("Inntektsrapportering_template.xlsx")
sheet_income_report = income_report.active

def check_sales_employee(goods):
    for row_payment in range(1, sheet_sales_employee.max_row + 1):
        for column_payment in "A":
            cell_name_sales = "{}{}".format(column_payment, row_payment)
            goods_sold_employee = sheet_sales_employee[cell_name_sales].value
        if goods == goods_sold_employee:
            amount_sold_location = "{}{}".format("C", row_payment)
            amount_sold_employee = sheet_sales_employee[amount_sold_location].value
            return goods_sold_employee, amount_sold_employee
    if goods != goods_sold_employee:
        amount_sold_employee = "0"
        return goods, amount_sold_employee

def check_sales_customer(goods):
    for row_payment in range(1, sheet_sales_customer.max_row + 1):
        for column_payment in "A":
            cell_name_sales = "{}{}".format(column_payment, row_payment)
            goods_sold_customer = sheet_sales_customer[cell_name_sales].value
        if goods == goods_sold_customer:
            amount_sold_location = "{}{}".format("C", row_payment)
            amount_sold_customer = sheet_sales_customer[amount_sold_location].value
            return goods_sold_customer, amount_sold_customer
    if goods != goods_sold_customer:
        amount_sold_customer = "0"
        return goods, amount_sold_customer

goods_name = []

for row in range(1, sheet_goods.max_row + 1):
    cell_name = "{}{}".format("A", row)
    goods_name.append(sheet_goods[cell_name].value)

for i in goods_name:
    sold_customer_goods, sold_customer_amount = check_sales_customer(i)
    for j in range(1, sheet_income_report.max_row + 1):
        cell_name_report = "{}{}".format("A", j)
        goods_name_report = sheet_income_report[cell_name_report].value
        if sold_customer_goods == goods_name_report:
            sheet_income_report.cell(row=int(j), column=int(8), value=int(sold_customer_amount))

for i in goods_name:
    sold_employee_goods, sold_employee_amount = check_sales_employee(i)
    for j in range(1, sheet_income_report.max_row + 1):
        cell_name_report = "{}{}".format("A", j)
        goods_name_report = sheet_income_report[cell_name_report].value
        if sold_employee_goods == goods_name_report:
            sheet_income_report.cell(row=int(j), column=int(9), value=int(sold_employee_amount))


regex = re.compile(r"(\d{1,2}.\d{1,2}.\d{1,4})")
matchArray = regex.findall(employee_sales)
date1 = matchArray[0]
date2 = matchArray[1]
first_day_previous_month  = datetime.strptime(date1, "%d.%m.%Y").strftime('%d.%m.%Y')
last_day_previous_month  = datetime.strptime(date2, "%d.%m.%Y").strftime('%d.%m.%Y')
month = int(datetime.strptime(date2, "%d.%m.%Y").strftime('%m'))
month += 1
month = str(month)

last_month = datetime.strptime(date2, "%d.%m.%Y").strftime('%B')
current_month = datetime.strptime(month, "%m").strftime("%B")
current_year = datetime.strptime(date2, "%d.%m.%Y").strftime("%Y")
first_day_current_month = (datetime.strptime(first_day_previous_month, '%d.%m.%Y') + relativedelta(months=1)).strftime('%d.%m.%Y')
last_day_current_month = (datetime.strptime(last_day_previous_month, '%d.%m.%Y') + relativedelta(months=1)).strftime('%d.%m.%Y')
month_year = current_month + " " + current_year

sheet_income_report.cell(row=1, column=6, value=month_year)
sheet_income_report.cell(row=2, column=6, value=("Telling \n" + last_day_current_month))
sheet_income_report.cell(row=2, column=7, value=("Lagerverdi eks. MVA \n" + last_day_current_month))
sheet_income_report.cell(row=2, column=8, value=("Salg \n" + last_month))
sheet_income_report.cell(row=2, column=9, value=("Ansattsalg \n" + last_month))
sheet_income_report.cell(row=2, column=10, value=("Salgsum \n" + last_month))
sheet_income_report.cell(row=21, column=11, value=("Vipps inn på konto (" + first_day_current_month + ")"))
sheet_income_report.cell(row=23, column=1, value=("Treningsinntekter fra Wodify (" + first_day_current_month + ") - (" + last_day_current_month + ")"))


drinks_nocco = 0
drinks_vann = 0
drinks_powerade = 0
drinks_barebells = 0
drinks_yt = 0
snacks_proteinchips = 0
snacks_chocolate = 0
misc_kalkbit = 0
misc_sportstape = 0
misc_wire = 0
misc_dropin = 0
clothes_tshirt = 0
clothes_hoodie = 0
clothes_babybody = 0
clothes_longsleeve = 0
clothes_tanktop = 0


for i in range(1, sheet_varetelling.max_row + 1):
    cell_name = "{}{}".format("A", i)
    brand = sheet_varetelling[cell_name].value
    if brand != None and brand == "Nocco":
        cell_name = "{}{}".format("C", i)
        drinks_nocco += sheet_varetelling[cell_name].value
    elif brand != None and brand == "Vann":
        cell_name = "{}{}".format("C", i)
        drinks_vann += sheet_varetelling[cell_name].value
    elif brand != None and brand == "Powerade":
        cell_name = "{}{}".format("C", i)
        drinks_powerade += sheet_varetelling[cell_name].value
    elif brand != None and brand in "Barebells Milkshake":
        cell_name = "{}{}".format("C", i)
        drinks_barebells += sheet_varetelling[cell_name].value
    elif brand != None and brand in "Yt Restitusjonsdrikk":
        cell_name = "{}{}".format("C", i)
        drinks_yt += sheet_varetelling[cell_name].value
    elif brand != None and brand == "Proteinchips":
        cell_name = "{}{}".format("C", i)
        snacks_proteinchips += sheet_varetelling[cell_name].value
    elif brand != None and brand in "Snickers, Mars, Bounty og M&M":
        cell_name = "{}{}".format("C", i)
        snacks_chocolate += sheet_varetelling[cell_name].value
    elif brand != None and brand == "Kalkbit":
        cell_name = "{}{}".format("C", i)
        misc_kalkbit += sheet_varetelling[cell_name].value
    elif brand != None and brand == "Sportstape":
        cell_name = "{}{}".format("C", i)
        misc_sportstape += sheet_varetelling[cell_name].value
    elif brand != None and brand == "Stålwire til hoppetau":
        cell_name = "{}{}".format("C", i)
        misc_wire += sheet_varetelling[cell_name].value
    elif brand != None and brand == "CFG T-Shirt":
        cell_name = "{}{}".format("C", i)
        clothes_tshirt += sheet_varetelling[cell_name].value
    elif brand != None and brand == "CFG Hoodie":
        cell_name = "{}{}".format("C", i)
        clothes_hoodie += sheet_varetelling[cell_name].value
    elif brand != None and brand == "CFG Baby Body":
        cell_name = "{}{}".format("C", i)
        clothes_babybody += sheet_varetelling[cell_name].value
    elif brand != None and brand == "CFG Longsleeve / Baseball Tee":
        cell_name = "{}{}".format("C", i)
        clothes_longsleeve += sheet_varetelling[cell_name].value
    elif brand != None and brand == "CFG Tanktop / Cropped Tee":
        cell_name = "{}{}".format("C", i)
        clothes_tanktop += sheet_varetelling[cell_name].value
    elif brand != None and brand == "Drop-in Trening":
        cell_name = "{}{}".format("C", i)
        misc_dropin += sheet_varetelling[cell_name].value


sheet_income_report.cell(row=int(3), column=int(6), value=int(drinks_nocco))
sheet_income_report.cell(row=int(4), column=int(6), value=int(drinks_vann))
sheet_income_report.cell(row=int(5), column=int(6), value=int(drinks_powerade))
sheet_income_report.cell(row=int(6), column=int(6), value=int(drinks_barebells))
sheet_income_report.cell(row=int(7), column=int(6), value=int(drinks_yt))
sheet_income_report.cell(row=int(8), column=int(6), value=int(snacks_proteinchips))
sheet_income_report.cell(row=int(9), column=int(6), value=int(snacks_chocolate))
sheet_income_report.cell(row=int(10), column=int(6), value=int(misc_kalkbit))
sheet_income_report.cell(row=int(11), column=int(6), value=int(misc_sportstape))
sheet_income_report.cell(row=int(12), column=int(6), value=int(misc_wire))
sheet_income_report.cell(row=int(13), column=int(6), value=int(clothes_tshirt))
sheet_income_report.cell(row=int(14), column=int(6), value=int(clothes_hoodie))
sheet_income_report.cell(row=int(15), column=int(6), value=int(clothes_babybody))
sheet_income_report.cell(row=int(16), column=int(6), value=int(clothes_longsleeve))
sheet_income_report.cell(row=int(17), column=int(6), value=int(clothes_tanktop))
sheet_income_report.cell(row=int(18), column=int(6), value=int(misc_dropin))